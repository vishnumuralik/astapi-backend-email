from fastapi import FastAPI, Form, File, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from email.message import EmailMessage
import aiosmtplib
import os
from openpyxl import Workbook, load_workbook
import mimetypes
from dotenv import load_dotenv

# Load environment variables
load_dotenv(dotenv_path=".env")

app = FastAPI()

# CORS setup (match your frontend URL here)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://user-input-healthproof.vercel.app"],  # ✅ no trailing slash
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
# Secure config from .env
GMAIL_USER = os.getenv("GMAIL_USER")
GMAIL_APP_PASSWORD = os.getenv("GMAIL_APP_PASSWORD")
EXCEL_FILE = os.getenv("EXCEL_FILE", "data.xlsx")  # default if not defined

@app.post("/send-email")
async def send_email(
    name: str = Form(...),
    totalExperience: str = Form(...),
    previousCompany: str = Form(...),
    domainSkill: str = Form(...),
    hobbies: str = Form(...),
    officeLocation: str = Form(...),
    photo: UploadFile = File(None),
):
    # Step 1: Save to Excel
    headers = ["Name", "Total Experience", "Previous Company", "Domain / Skill", "Hobbies", "Office Location"]
    row = [name, totalExperience, previousCompany, domainSkill, hobbies, officeLocation]

    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
    else:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

    ws.append(row)
    wb.save(EXCEL_FILE)

    # Step 2: Email content
    message = EmailMessage()
    message["From"] = GMAIL_USER
    message["To"] = GMAIL_USER
    message["Subject"] = f"🎉 Welcome to Team – {name}"

    # Plain text
    message.set_content(
        f"""
{name} has submitted their professional details.

They have {totalExperience} years of experience.
They previously worked at {previousCompany}.
Their domain/skill is {domainSkill}.
Their hobbies include {hobbies}.
Preferred office location is {officeLocation}.
"""
    )

    # HTML + Image handling
    photo_bytes = None
    photo_type = None
    cid = "profilephoto@act"
    cid_img_tag = ""

    if photo and photo.filename:
        content_type = photo.content_type or "image/jpeg"
        if content_type not in ["image/jpeg", "image/png", "image/jpg"]:
            raise HTTPException(status_code=400, detail="Unsupported image format.")

        photo_bytes = await photo.read()
        if not photo_bytes:
            raise HTTPException(status_code=400, detail="Uploaded image is empty.")

        photo_type = mimetypes.guess_extension(content_type).replace(".", "") or "jpeg"

        cid_img_tag = f"""
        <div style="text-align: center; margin-top: 20px;">
          <img src="cid:{cid}" alt="Profile Photo"
               style="border-radius: 50%; width: 150px; height: 150px;
                      object-fit: cover; border: 2px solid #ccc;
                      box-shadow: 0 0 10px rgba(0,0,0,0.1);" />
        </div>
        """

    # Add HTML alternative
    message.add_alternative(f"""
    <html>
      <body style="font-family: Arial, sans-serif; padding: 20px; background-color: #f4f4f4;">
        <div style="background: white; padding: 20px; border-radius: 10px; max-width: 600px; margin: auto;">
          <h2 style="color: #003366;">🎉 Welcome to ACT – {name}</h2>
          <p style="font-size: 16px; line-height: 1.7;">
            <strong>{name}</strong> has submitted their professional details.<br><br>
            They have <strong>{totalExperience}</strong> years of experience.<br>
            They previously worked at <strong>{previousCompany}</strong>.<br>
            Their primary domain or skill is <strong>{domainSkill}</strong>.<br>
            They enjoy <strong>{hobbies}</strong> as hobbies.<br>
            Their preferred office location is <strong>{officeLocation}</strong>.
          </p>
          {cid_img_tag}
          <p style="margin-top: 30px;">Best regards,<br><strong>Professional Details System</strong></p>
        </div>
      </body>
    </html>
    """, subtype="html")

    # Attach Excel
    with open(EXCEL_FILE, "rb") as f:
        message.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="data.xlsx"
        )

    # Attach photo (CID image)
    if photo_bytes and photo_type:
        message.add_attachment(
            photo_bytes,
            maintype="image",
            subtype=photo_type,
            filename=photo.filename,
            cid=cid
        )

    # Send email
    try:
        await aiosmtplib.send(
            message,
            hostname="smtp.gmail.com",
            port=587,
            start_tls=True,
            username=GMAIL_USER,
            password=GMAIL_APP_PASSWORD,
        )
        return {"message": "✅ Email sent successfully with narrative details! Thanks for participating"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Email failed: {e}")
